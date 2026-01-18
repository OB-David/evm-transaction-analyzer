import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Dict

def generate_table_excel(self, output_path: str = "transaction_table.xlsx") -> None:
    """
    基于维护的table列表生成Excel文件，为不同地址自动分配不同背景色
    :param output_path: Excel输出路径
    """
    if not self.table:
        print("警告：table列表为空，生成空Excel")
        return
    
    # 准备15种颜色
    COLOR_PALETTE = [
        "E8F4FD",  # 浅蓝（浅）
        "F0F8E8",  # 浅绿（浅）
        "FDF2E8",  # 浅橙（浅）
        "F8E8FD",  # 浅紫（浅）
        "E8FDF0",  # 浅青（浅）
        "FDE8E8",  # 浅红（浅）
        "F5F2E8",  # 浅黄（浅）
        "F8E8F0",  # 浅粉（浅）
        "B3D9F2",  # 天蓝（深）
        "C9E4B3",  # 草绿（深）
        "F2D0B3",  # 橙黄（深）
        "E0B3F2",  # 淡紫（深）
        "B3F2CC",  # 青绿（深）
        "F2B3B3",  # 淡红（深）
        "E8D9B3",  # 米黄（深）
    ]
    # 地址-颜色映射字典（确保同一地址始终用同一种颜色）
    address_color_map: Dict[str, PatternFill] = {}
    color_index = 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaction Table"
    
    # 定义表头（与table字段对应）
    headers = ["pc", "op", "from", "to", "token", "balance/amount"]
    # 设置表头样式
    header_font = Font(bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    # 表头背景色（突出表头）
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = header_fill  # 表头添加浅灰色背景
    
    # ===================== 修改：写入数据并染色 =====================
    for row_idx, row_data in enumerate(self.table, 2):
        for col_idx, header in enumerate(headers, 1):
            cell_value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = Alignment(horizontal="center", vertical="center")  # 内容居中
            
            # 只为from/to列的地址添加背景色
            if header in ["from", "to"] and cell_value:  # 非空地址才染色
                # 如果地址未分配颜色，自动分配（超过15种则循环）
                if cell_value not in address_color_map:
                    color_hex = COLOR_PALETTE[color_index % len(COLOR_PALETTE)]
                    address_color_map[cell_value] = PatternFill(
                        start_color=color_hex,
                        end_color=color_hex,
                        fill_type="solid"
                    )
                    color_index += 1
                # 应用颜色
                cell.fill = address_color_map[cell_value]
    
    # ===================== 修复：列宽匹配（原代码多了1个值） =====================
    col_widths = [10, 10, 45, 45, 45, 25]  # 对应6个表头：pc,op,from,to,token,balance/amount
    for col_idx, width in enumerate(col_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
    
    # 保存Excel
    wb.save(output_path)
    print(f"✅ 表格数据已生成至Excel：{output_path}")
    print(f"📊 共记录 {len(self.table)} 条数据（SSTORE/SLOAD/CALL）")
    print(f"🎨 共为 {len(address_color_map)} 个不同地址分配了颜色")